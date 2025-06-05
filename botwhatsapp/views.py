from django.shortcuts import render
from .models import AvaliacaoAtendimento
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.timezone import make_naive
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
# Create your views here.

def avaliacoes_atendimento_whatsapp(request):

    avaliacoes = AvaliacaoAtendimento.objects.all().order_by('-data')

    context = {
        'avaliacoes': avaliacoes
    }

    return render(request, 'avaliacoes_atendimento.html', context)

def avaliacoes_atendimento_whatsapp_exportar(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Avaliações de Atendimento"

    headers = ['ID', 'Telefone', 'Nome', 'Avaliação', 'Detalhamento', 'Data']
    worksheet.append(headers)

    # Estilização do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Ajuste de largura das colunas
    column_widths = [8, 18, 20, 15, 40, 20]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[chr(64 + i)].width = width

    for row_idx, avaliacao in enumerate(AvaliacaoAtendimento.objects.all().order_by('-data'), start=2):
        data_formatada = make_naive(avaliacao.data).strftime('%d/%m/%Y %H:%M')
        worksheet.append([
            avaliacao.id,
            avaliacao.telefone,
            getattr(avaliacao, 'nome', ''),
            dict(AvaliacaoAtendimento._meta.get_field('avaliacao').choices).get(avaliacao.avaliacao, avaliacao.avaliacao),
            getattr(avaliacao, 'detalhamento', ''),
            data_formatada
        ])
        # Adiciona borda nas células da linha
        for col_num in range(1, 7):
            worksheet.cell(row=row_idx, column=col_num).border = thin_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="avaliacoes_atendimento.xlsx"'
    workbook.save(response)
    return response