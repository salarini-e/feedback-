from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Avg, Count
from .forms import FeedbackForm, PesquisaSatisfacaoForm
from servicos.models import Local_de_atendimento
from .functions import get_feedback_distribution, get_frequencies, get_valores_termometro
from .models import Feedback
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

def excecoes(request, hash):
    if hash == 'cc10d409909d729cd065b242bbe22b20' or hash == '3cca8f55a6ef272be6009a887bd83ae6':
        return pesquisa_satisfacao(request, hash)
    # elif hash == 'eaf6b20893af78a08f6f426a99f045b1':
    #     return painel_feedback_ti(request, hash)
    return None

def feedback(request, hash):
    # Verifica se o hash corresponde a uma exceção
    resposta = excecoes(request, hash)
    if resposta:
        return resposta  # Redireciona para a pesquisa de satisfação

    print('passou aqui')
    local = get_object_or_404(Local_de_atendimento, hash=hash)
    if request.method == 'POST':
        print(request.POST)
        data = {
            'local': local.id,
            'ambiente': request.POST.get('ambiente'),
            'tempo_espera': request.POST.get('tempo'),
            'satisfacao': request.POST.get('satisfacao'),
            'atendimento': request.POST.get('atendimento'),
            'sugestoes': request.POST.get('sugestao'),
            'incluir_contato': bool(request.POST.get('resposta')),
        }
        if bool(request.POST.get('resposta')):
            data['contato_nome'] = request.POST.get('contato_nome')
            data['contato_telefoe'] = request.POST.get('contato_telefone')
        form = FeedbackForm(data)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.local = local
            feedback.save()                        
            return render(request, 'feedback_success.html', {'local': local})
        else:
            messages.error(request, "Por favor, certifique-se de que todas as perguntas foram respondidas.")
    else:
        form = FeedbackForm()

    context = {
        'local': local,
        'form': form
    }    
    return render(request, 'feedback.html', context)

#PESQUISA DE SATISFAÇÃO
def pesquisa_satisfacao(request, hash):
    local = get_object_or_404(Local_de_atendimento, hash=hash)
    if request.method == 'POST':
        print(request.POST)
        form = PesquisaSatisfacaoForm(request.POST)
        form.fields['local'].initial = local.id  # Define o local automaticamente
        if form.is_valid():
            pesquisa = form.save(commit=False)
            pesquisa.local = local
            pesquisa.save()
            return render(request, 'feedback_success.html', {'local': local})
        else:
            messages.error(request, "Por favor, corrija os erros no formulário.")
            print(form.errors)
    else:
        form = PesquisaSatisfacaoForm()
    context = {'form': form, 'local': local}
    if form.errors:
        context['data']= request.POST
    return render(request, 'feedback pesquisa_satisfacao.html', context)

def painel_excecoes(request, hash):
    
    if hash == 'cc10d409909d729cd065b242bbe22b20' or hash == '3cca8f55a6ef272be6009a887bd83ae6':
        return painel_pesquisa_satisfacao(request, hash)
    elif hash == 'b0e38ec238850c01bb2beb853483ea08' or hash == 'eaf6b20893af78a08f6f426a99f045b1':
        return painel_feedback_ti(request, hash)
    return False


def painel_feedback(request, hash):
    # Verifica se o hash corresponde a uma exceção
    resposta = painel_excecoes(request, hash)
    if resposta:
        return resposta
    
    print('passou aqui')
    local = get_object_or_404(Local_de_atendimento, hash=hash)
    # feedbacks = local.feedbacks.all()    
    feedbacks = Feedback.objects.filter(local=local).order_by('-dt_inclusao')
    negative_feedbacks = feedbacks.filter(satisfacao__lt=3)
    feedbacks_with_suggestions = feedbacks.exclude(sugestoes__isnull=True).exclude(sugestoes__exact="").exclude(sugestoes__iexact="n/h")  # Exclui 'n/h'

    valores_satisfacao_geral = get_valores_termometro(feedbacks)
    summary = {
        
        'ambiente_dist': get_feedback_distribution(feedbacks, 'ambiente'),
        'tempo_espera_dist': get_feedback_distribution(feedbacks, 'tempo_espera'),
        'satisfacao_dist': get_feedback_distribution(feedbacks, 'satisfacao'),  
        'atendimento_dist': get_feedback_distribution(feedbacks, 'atendimento'),
        'total_feedbacks': feedbacks.count(),
        'negative_feedbacks': feedbacks.filter(satisfacao__lt=3).count(),
    }

    # Add moda and mediana to each metric
    contexto_estatisticas = {
        'ambiente': get_frequencies(feedbacks, 'ambiente'),
        'tempo_espera': get_frequencies(feedbacks, 'tempo_espera'),
        'satisfacao': get_frequencies(feedbacks, 'satisfacao'),
        'atendimento': get_frequencies(feedbacks, 'atendimento'),
    }

    summary_with_flags = [
        {
            'metric': key.replace('_dist', '').capitalize(),
            'is_distribution': key.endswith('_dist'),
            'data': value,
            'moda': contexto_estatisticas[key.replace('_dist', '')]['moda'],
            'media': contexto_estatisticas[key.replace('_dist', '')]['media']
        }
        for key, value in summary.items() if key.endswith('_dist')
    ]

    # Pass non-distribution keys separately
    non_distribution_summary = {
        'total_feedbacks': summary['total_feedbacks'],
        'negative_feedbacks': summary['negative_feedbacks'],
    }

    context = {
        'local': local,
        'feedbacks': feedbacks,
        'feedbacks_with_suggestions': feedbacks_with_suggestions,  # Atualizado para excluir 'n/h'
        'negative_feedbacks': negative_feedbacks,
        'summary': summary_with_flags,
        'non_distribution_summary': non_distribution_summary,
        'estatisticas': contexto_estatisticas,
        'valores_satisfacao_geral': valores_satisfacao_geral
    }

    # if local.nome == 'SUBSECRETARIA DE TI':
    #     return render(request, 'painel_fedback_ti.html', context)
    return render(request, 'painel_fedback.html', context)


def painel_feedback_ti(request, hash):
    print('passou aqui no painel_feedback_ti')
    # Verifica se o hash corresponde a uma exceção
    local = get_object_or_404(Local_de_atendimento, hash=hash)
    feedbacks = Feedback.objects.filter(local=local).order_by('-dt_inclusao')
    print('total de feedbacks:', feedbacks.count())
    negative_feedbacks = feedbacks.filter(satisfacao__lt=3)
    feedbacks_with_suggestions = feedbacks.exclude(sugestoes__isnull=True).exclude(sugestoes__exact="").exclude(sugestoes__iexact="n/h")  # Exclui 'n/h'

    valores_satisfacao_geral = get_valores_termometro(feedbacks)
    summary = {
        
        'ambiente_dist': get_feedback_distribution(feedbacks, 'ambiente'),
        'tempo_espera_dist': get_feedback_distribution(feedbacks, 'tempo_espera'),
        'satisfacao_dist': get_feedback_distribution(feedbacks, 'satisfacao'),  
        'atendimento_dist': get_feedback_distribution(feedbacks, 'atendimento'),
        'total_feedbacks': feedbacks.count(),
        'negative_feedbacks': feedbacks.filter(satisfacao__lt=3).count(),
    }

    # Add moda and mediana to each metric
    contexto_estatisticas = {
        'ambiente': get_frequencies(feedbacks, 'ambiente'),
        'tempo_espera': get_frequencies(feedbacks, 'tempo_espera'),
        'satisfacao': get_frequencies(feedbacks, 'satisfacao'),
        'atendimento': get_frequencies(feedbacks, 'atendimento'),
    }

    summary_with_flags = [
        {
            'metric': key.replace('_dist', '').capitalize(),
            'is_distribution': key.endswith('_dist'),
            'data': value,
            'moda': contexto_estatisticas[key.replace('_dist', '')]['moda'],
            'media': contexto_estatisticas[key.replace('_dist', '')]['media']
        }
        for key, value in summary.items() if key.endswith('_dist')
    ]

    # Pass non-distribution keys separately
    non_distribution_summary = {
        'total_feedbacks': summary['total_feedbacks'],
        'negative_feedbacks': summary['negative_feedbacks'],
    }

    context = {
        'local': local,
        'feedbacks': feedbacks,
        'feedbacks_with_suggestions': feedbacks_with_suggestions,  # Atualizado para excluir 'n/h'
        'negative_feedbacks': negative_feedbacks,
        'summary': summary_with_flags,
        'non_distribution_summary': non_distribution_summary,
        'estatisticas': contexto_estatisticas,
        'valores_satisfacao_geral': valores_satisfacao_geral
    }

    return render(request, 'painel_fedback_ti.html', context)

from .models import PesquisaSatisfacao
def painel_pesquisa_satisfacao(request, hash):
    local = get_object_or_404(Local_de_atendimento, hash=hash)
    pesquisas = PesquisaSatisfacao.objects.filter(local=local).order_by('-dt_inclusao')

    # Campos para distribuição (exemplo: internet_item1, telefonia_item1, etc)
    campos_distribuicao = [
        ('Internet', 'internet_item1'),
        ('Telefonia', 'telefonia_item1'),
        ('Impressora', 'impressora_item1'),
        ('Sistema Gestão', 'sistema_gestao_item4'),
        ('Processo Digital', 'processo_item1'),
        ('Tramitação', 'tramitacao_item1'),
        ('Suporte', 'suporte_item1'),
    ]

    summary = []
    for nome, campo in campos_distribuicao:
        # Calcula distribuição
        dist = {}
        total = pesquisas.count()
        for nota in range(1, 6):
            count = pesquisas.filter(**{campo: nota}).count()
            percent = round((count / total) * 100, 1) if total else 0
            dist[nota] = {'count': count, 'percent': percent}
        # Moda
        moda = None
        moda_qtd = 0
        for nota in range(1, 6):
            if dist[nota]['count'] > moda_qtd:
                moda = nota
                moda_qtd = dist[nota]['count']
        # Mediana
        notas = list(pesquisas.values_list(campo, flat=True))
        notas_sorted = sorted(notas)
        n = len(notas_sorted)
        if n == 0:
            mediana = None
        elif n % 2 == 1:
            mediana = notas_sorted[n // 2]
        else:
            mediana = (notas_sorted[n // 2 - 1] + notas_sorted[n // 2]) / 2
        summary.append({
            'metric': nome,
            'is_distribution': True,
            'data': dist,
            'moda': moda,
            'mediana': mediana
        })

    context = {
        'local': local,
        'pesquisas': pesquisas,
        'summary': summary,
    }
    return render(request, 'painel_fedback_pesquisa_satisfacao.html', context)

def exportar_pesquisas_excel(request, hash):
    from .models import PesquisaSatisfacao
    local = get_object_or_404(Local_de_atendimento, hash=hash)
    pesquisas = PesquisaSatisfacao.objects.filter(local=local).order_by('-dt_inclusao')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pesquisas"

    # Cabeçalhos
    headers = [
        "ID", "Data", "Secretaria",
        "Internet 1", "Internet 2", "Internet 3", "Internet 3 Porquê", "Internet 4", "Internet 5",
        "Telefonia 1", "Telefonia 2", "Telefonia 2 Porquê", "Telefonia 3", "Telefonia 4", "Telefonia 4 Porquê", "Telefonia 5", "Telefonia 6",
        "Impressora 1", "Impressora 2", "Impressora 3", "Impressora 4", "Impressora 5", "Impressora 5 Porquê", "Impressora 6",
        "Sistema Gestão 1", "Sistema Gestão 2", "Sistema Gestão 2 Porquê", "Sistema Gestão 3", "Sistema Gestão 3 Porquê", "Sistema Gestão 4", "Sistema Gestão 5",
        "Processo 1", "Processo 2", "Processo 3",
        "Tramitação 1", "Tramitação 2", "Tramitação 3", "Tramitação 4",
        "Suporte 1", "Suporte 2", "Suporte 3", "Suporte 4.2",
        "Resposta", "Contato Nome", "Contato Telefone"
    ]
    ws.append(headers)

    for pesquisa in pesquisas:
        ws.append([
            pesquisa.id,
            pesquisa.dt_inclusao.strftime("%d/%m/%Y %H:%M"),
            pesquisa.secretaria,
            pesquisa.internet_item1,
            pesquisa.internet_item2,
            pesquisa.internet_item3,
            pesquisa.internet_item3_porque,
            pesquisa.internet_item4,
            pesquisa.internet_item5,
            pesquisa.telefonia_item1,
            pesquisa.telefonia_item2,
            pesquisa.telefonia_item2_porque,
            pesquisa.telefonia_item3,
            pesquisa.suporte_item4,
            pesquisa.suporte_item4_porque,
            pesquisa.telefonia_item5,
            pesquisa.telefonia_item6,
            pesquisa.impressora_item1,
            pesquisa.impressora_item2,
            pesquisa.impressora_item3,
            pesquisa.impressora_item4,
            pesquisa.impressora_item5,
            pesquisa.impressora_item5_porque,
            pesquisa.impressora_item6,
            pesquisa.sistema_gestao_item1,
            pesquisa.sistema_gestao_item2,
            pesquisa.sistema_gestao_item2_porque,
            pesquisa.sistema_gestao_item3,
            pesquisa.sistema_gestao_item3porque,
            pesquisa.sistema_gestao_item4,
            pesquisa.sistema_gestao_item5,
            pesquisa.processo_item1,
            pesquisa.processo_item2,
            pesquisa.processo_item3,
            pesquisa.tramitacao_item1,
            pesquisa.tramitacao_item2,
            pesquisa.tramitacao_item3,
            pesquisa.tramitacao_item4,
            pesquisa.suporte_item1,
            pesquisa.suporte_item2,
            pesquisa.suporte_item3,
            pesquisa.suporte_item42,
            pesquisa.resposta,
            pesquisa.contato_nome,
            pesquisa.contato_telefone,
        ])

    # Ajusta largura das colunas
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(max_length + 2, 40))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=pesquisas_{local.hash}.xlsx'
    wb.save(response)
    return response